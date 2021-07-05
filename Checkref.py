import os as sys
import webbrowser
import openpyxl as xl


class speedGrade:

    def __init__(self, xlIntake, docCache):
        # Initialize variables
        self.xlIntake = xlIntake
        self.docCache = docCache
        self.problemChecks = None
        self.mySheet = None
        self.probHead = []
        self.names = []

        # Send object to processing step
        self.findProb()
        self.getHeaders()
        self.getNames()

    def findProb(self):
        # This function finds the problems in the spreadsheet and sorts them into groups
        self.intake = xl.load_workbook(self.xlIntake)

        mySheet = self.intake.active
        self.mySheet = mySheet

        # problemChecks stores the beginning and end of each problem
        self.problemChecks = [[], []]

        # Determine Number of Problems and their locations
        for cols in range(1, mySheet.max_column):
            if 'Total' in mySheet.cell(1, cols).value:
                self.problemChecks[1].append(cols - 1)
                for col in range(cols, -1, -1):
                    if mySheet.cell(1, col).value == "Email" or "Comments" in mySheet.cell(1, col).value:
                        self.problemChecks[0].append(col + 1)
                        break
            # Finds and stores the location of searchable names
            if mySheet.cell(1, cols).value == 'Sortable name':
                self.nameLoc = cols

    # Gets a list of names
    def getNames(self):
        for name in range(2, self.mySheet.max_row + 1):
            self.names.append(self.mySheet.cell(name, 2).value)

    # Finds the problem headers and returns them in list form
    def getHeaders(self):
        for prob in range(len(self.problemChecks[0])):
            for rng in range(self.problemChecks[0][prob], self.problemChecks[1][prob] + 1):
                self.probHead.append(self.mySheet.cell(1, rng).value)

    # zeroize finds students who did not submit work and zeroizes the grade
    def zeroize(self):
        docCache = self.docCache
        mySheet = self.mySheet
        problemChecks = self.problemChecks
        self.absCount = 0

        docs = sys.listdir("{}".format(docCache))
        for rows in range(2, mySheet.max_row + 1):
            personFound = False
            name = mySheet.cell(rows, self.nameLoc).value
            name = name.replace(" ", "").replace(",", "").lower()
            lastOnly = mySheet.cell(rows, self.nameLoc).value
            lastOnly = lastOnly.split(',')
            lastOnly = lastOnly[0].lower()
            for files in range(len(docs)):
                if name in docs[files]:
                    personFound = True
                    break
                elif lastOnly in docs[files]:
                    personFound = True
                    break
            # If person is not found among files then set all grades to 0
            if personFound == False:
                self.absCount += 1
                for probs in range(len(problemChecks[0])):
                    for sections in range(problemChecks[0][probs], (problemChecks[1][probs]) + 1):
                        mySheet.cell(row=rows, column=sections).value = 0

    # Gets the student's PDF file
    def pullPDF(self, index):
        docCache = self.docCache
        mySheet = self.mySheet

        docs = sys.listdir(docCache)
        name = mySheet.cell(index + 2, self.nameLoc).value
        name = name.replace(" ", "").replace(",", "").lower()
        lastOnly = mySheet.cell(index + 2, self.nameLoc).value
        lastOnly = lastOnly.split(',')
        lastOnly = lastOnly[0].lower()
        for files in range(len(docs)):
            if name in docs[files]:
                webbrowser.open_new('"{}/{}"'.format(self.docCache, docs[files]))
                return True
                break
            elif lastOnly in docs[files]:
                webbrowser.open_new('"{}/{}"'.format(self.docCache, docs[files]))
                return True
                break
        return False

    # Takes in the index value for a student and returns the scores in the spreadsheet
    def pullScores(self, indexLoc):
        scores = []
        problemChecks = self.problemChecks
        for probs in range(len(problemChecks[0])):
            for sections in range(problemChecks[0][probs], (problemChecks[1][probs]) + 1):
                scores.append(self.mySheet.cell(row=indexLoc + 2, column=sections).value)
        return scores

    # Saves the local excel data to the excel spreadsheet
    def save(self):
        # removes the existing spreadsheet and replaces it with an updated copy
        sys.remove(self.xlIntake)
        self.intake.save('{}'.format(self.xlIntake))


def main():
    speedy = speedGrade('Homework6_Gradesheet.xlsx', 'submissions')
    speedy.zeroize()
    speedy.pullPDF(0)
    speedy.save()

if __name__ == "__main__":
    main()